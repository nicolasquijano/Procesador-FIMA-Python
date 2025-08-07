"""
Exportador Excel con formato profesional y múltiples hojas
"""
import os
from decimal import Decimal
from datetime import datetime
from typing import List, Dict, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from config import EXPORT_DIR

class ExcelExporter:
    """Exportador Excel con formato profesional"""
    
    def __init__(self):
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.right_alignment = Alignment(horizontal='right', vertical='center')
    
    def _format_decimal_for_excel(self, value) -> float:
        """Convertir Decimal a float para Excel"""
        if isinstance(value, Decimal):
            return float(value)
        elif isinstance(value, str):
            try:
                return float(Decimal(value))
            except:
                return value
        return value
    
    def _prepare_operations_data(self, operations: List[Dict]) -> pd.DataFrame:
        """Preparar datos de operaciones para Excel"""
        if not operations:
            return pd.DataFrame()
        
        # Convertir operaciones a formato Excel
        excel_data = []
        for op in operations:
            excel_row = {
                'ID': op.get('id', ''),
                'Fecha': op.get('date', ''),
                'Tipo Operación': op.get('operation_type', ''),
                'Fondo': op.get('fund_name', ''),
                'Tipo Fondo': op.get('fund_type', ''),
                'Cuotas': self._format_decimal_for_excel(op.get('quantity', 0)),
                'Valor Unitario': self._format_decimal_for_excel(op.get('unit_value', 0)),
                'Monto Total': self._format_decimal_for_excel(op.get('total_amount', 0)),
                'Descripción': op.get('description', ''),
                'Fuente PDF': op.get('pdf_source', ''),
                'Fecha Creación': op.get('created_at', '')
            }
            excel_data.append(excel_row)
        
        return pd.DataFrame(excel_data)
    
    def _prepare_positions_data(self, positions: List[Dict]) -> pd.DataFrame:
        """Preparar datos de posiciones para Excel"""
        if not positions:
            return pd.DataFrame()
        
        excel_data = []
        for pos in positions:
            excel_row = {
                'ID': pos.get('id', ''),
                'Fondo': pos.get('fund_name', ''),
                'Tipo Fondo': pos.get('fund_type', ''),
                'Cuotas': self._format_decimal_for_excel(pos.get('quantity', 0)),
                'Valor Unitario': self._format_decimal_for_excel(pos.get('unit_value', 0)),
                'Valor Total': self._format_decimal_for_excel(pos.get('total_value', 0)),
                'Última Actualización': pos.get('last_updated', '')
            }
            excel_data.append(excel_row)
        
        return pd.DataFrame(excel_data)
    
    def _prepare_config_data(self, configs: List[Dict]) -> pd.DataFrame:
        """Preparar datos de configuración para Excel"""
        if not configs:
            return pd.DataFrame()
        
        excel_data = []
        for config in configs:
            excel_row = {
                'ID': config.get('id', ''),
                'Fondo': config.get('fund_name', ''),
                'Tipo Fondo': config.get('fund_type', ''),
                'Saldo Inicial': self._format_decimal_for_excel(config.get('initial_balance', 0)),
                'Activo': 'Sí' if config.get('active', 0) else 'No',
                'Fecha Creación': config.get('created_at', '')
            }
            excel_data.append(excel_row)
        
        return pd.DataFrame(excel_data)
    
    def _format_worksheet(self, ws, df: pd.DataFrame, title: str):
        """Aplicar formato profesional a una hoja"""
        if df.empty:
            ws['A1'] = f"No hay datos disponibles para {title}"
            ws['A1'].font = Font(bold=True, size=14)
            return
        
        # Agregar título
        ws.merge_cells(f'A1:{get_column_letter(len(df.columns))}1')
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=16, color="366092")
        ws['A1'].alignment = self.center_alignment
        
        # Agregar fecha de generación
        ws.merge_cells(f'A2:{get_column_letter(len(df.columns))}2')
        ws['A2'] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws['A2'].font = Font(italic=True, size=10)
        ws['A2'].alignment = self.center_alignment
        
        # Insertar datos (comenzando en fila 4)
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 4):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Aplicar formato a los encabezados
                if r_idx == 4:  # Fila de encabezados
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = self.center_alignment
                
                # Aplicar bordes
                cell.border = self.border
                
                # Alineación especial para números
                if isinstance(value, (int, float)) and r_idx > 4:
                    cell.alignment = self.right_alignment
                    # Formato de número para decimales
                    if isinstance(value, float):
                        cell.number_format = '#,##0.00000000'
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Congelar paneles (encabezados)
        ws.freeze_panes = 'A5'
    
    def _add_summary_sheet(self, wb: Workbook, operations: List[Dict], 
                          positions: List[Dict], stats: Dict):
        """Agregar hoja de resumen ejecutivo"""
        ws = wb.create_sheet("Resumen Ejecutivo", 0)
        
        # Título principal
        ws['A1'] = "PROCESADOR PDF FINANCIERO - RESUMEN EJECUTIVO"
        ws['A1'].font = Font(bold=True, size=18, color="366092")
        ws.merge_cells('A1:D1')
        ws['A1'].alignment = self.center_alignment
        
        # Fecha del reporte
        ws['A3'] = f"Fecha del Reporte: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws['A3'].font = Font(bold=True, size=12)
        
        # Estadísticas principales
        ws['A5'] = "ESTADÍSTICAS PRINCIPALES"
        ws['A5'].font = Font(bold=True, size=14, color="366092")
        
        stats_data = [
            ("Total de Operaciones:", stats.get('total_operations', 0)),
            ("Total de Posiciones:", stats.get('total_positions', 0)),
            ("Fondos Configurados:", stats.get('configured_funds', 0)),
            ("Valor Total del Portfolio:", f"${stats.get('total_portfolio_value', 0):,.2f}")
        ]
        
        for i, (label, value) in enumerate(stats_data, 6):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)
        
        # Resumen por tipo de fondo
        if positions:
            ws['A11'] = "RESUMEN POR TIPO DE FONDO"
            ws['A11'].font = Font(bold=True, size=14, color="366092")
            
            # Agrupar por tipo de fondo
            fund_summary = {}
            for pos in positions:
                fund_type = pos.get('fund_type', 'Sin Clasificar')
                if fund_type not in fund_summary:
                    fund_summary[fund_type] = {
                        'count': 0,
                        'total_value': Decimal('0')
                    }
                fund_summary[fund_type]['count'] += 1
                fund_summary[fund_type]['total_value'] += Decimal(str(pos.get('total_value', 0)))
            
            # Encabezados
            headers = ['Tipo de Fondo', 'Cantidad', 'Valor Total', '% del Portfolio']
            for i, header in enumerate(headers, 1):
                cell = ws.cell(row=13, column=i, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.center_alignment
                cell.border = self.border
            
            # Datos
            total_portfolio = stats.get('total_portfolio_value', Decimal('0'))
            for i, (fund_type, data) in enumerate(fund_summary.items(), 14):
                percentage = (data['total_value'] / total_portfolio * 100) if total_portfolio > 0 else 0
                
                row_data = [
                    fund_type,
                    data['count'],
                    float(data['total_value']),
                    f"{float(percentage):.2f}%"
                ]
                
                for j, value in enumerate(row_data, 1):
                    cell = ws.cell(row=i, column=j, value=value)
                    cell.border = self.border
                    if j in [2, 3]:  # Números
                        cell.alignment = self.right_alignment
                    if j == 3:  # Formato monetario
                        cell.number_format = '#,##0.00'
        
        # Ajustar anchos de columna
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def export_to_excel(self, operations: List[Dict], positions: List[Dict], 
                       configs: List[Dict], stats: Dict,
                       filename: Optional[str] = None,
                       fund_type_filter: Optional[str] = None) -> str:
        """Exportar datos a Excel con formato profesional"""
        
        # Filtrar operaciones por tipo de fondo si se especifica
        filtered_operations = operations
        if fund_type_filter:
            filtered_operations = [
                op for op in operations 
                if op.get('fund_type') == fund_type_filter
            ]
        
        # Generar nombre de archivo si no se proporciona
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filter_suffix = f"_{fund_type_filter}" if fund_type_filter else ""
            filename = f"reporte_financiero{filter_suffix}_{timestamp}.xlsx"
        
        filepath = os.path.join(EXPORT_DIR, filename)
        
        try:
            # Crear workbook
            wb = Workbook()
            
            # Eliminar hoja por defecto
            default_sheet = wb.active
            wb.remove(default_sheet)
            
            # Preparar DataFrames
            operations_df = self._prepare_operations_data(filtered_operations)
            positions_df = self._prepare_positions_data(positions)
            configs_df = self._prepare_config_data(configs)
            
            # Agregar hoja de resumen ejecutivo
            self._add_summary_sheet(wb, filtered_operations, positions, stats)
            
            # Hoja de operaciones
            if not operations_df.empty:
                ws_operations = wb.create_sheet("Operaciones")
                title = f"Operaciones"
                if fund_type_filter:
                    title += f" - {fund_type_filter}"
                self._format_worksheet(ws_operations, operations_df, title)
            
            # Hoja de posiciones
            if not positions_df.empty:
                ws_positions = wb.create_sheet("Posiciones Actuales")
                self._format_worksheet(ws_positions, positions_df, "Posiciones Actuales")
            
            # Hoja de configuración
            if not configs_df.empty:
                ws_configs = wb.create_sheet("Configuración de Fondos")
                self._format_worksheet(ws_configs, configs_df, "Configuración de Fondos")
            
            # Guardar archivo
            wb.save(filepath)
            
            return filepath
            
        except Exception as e:
            raise Exception(f"Error exportando a Excel: {str(e)}")
    
    def export_operations_csv(self, operations: List[Dict], 
                             filename: Optional[str] = None) -> str:
        """Exportar solo operaciones a CSV"""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"operaciones_{timestamp}.csv"
        
        filepath = os.path.join(EXPORT_DIR, filename)
        
        try:
            df = self._prepare_operations_data(operations)
            if df.empty:
                raise Exception("No hay operaciones para exportar")
            
            df.to_csv(filepath, index=False, encoding='utf-8')
            return filepath
            
        except Exception as e:
            raise Exception(f"Error exportando CSV: {str(e)}")
    
    def get_available_exports(self) -> List[Dict]:
        """Obtener lista de archivos exportados disponibles"""
        exports = []
        
        if not os.path.exists(EXPORT_DIR):
            return exports
        
        for filename in os.listdir(EXPORT_DIR):
            if filename.endswith(('.xlsx', '.csv')):
                filepath = os.path.join(EXPORT_DIR, filename)
                stat = os.stat(filepath)
                
                exports.append({
                    'filename': filename,
                    'filepath': filepath,
                    'size': stat.st_size,
                    'created': datetime.fromtimestamp(stat.st_ctime),
                    'modified': datetime.fromtimestamp(stat.st_mtime)
                })
        
        return sorted(exports, key=lambda x: x['modified'], reverse=True)
    
    def export_complete_report(self, operations: List[Dict], positions: List[Dict], 
                             fund_type_filter: Optional[str] = None) -> str:
        """Exportar reporte completo con todas las hojas"""
        # Obtener configuraciones y stats desde la base de datos
        from database import DatabaseManager
        db = DatabaseManager()
        
        configs = db.get_fund_configs()
        stats = db.get_database_stats()
        
        return self.export_to_excel(
            operations=operations,
            positions=positions,
            configs=configs,
            stats=stats,
            fund_type_filter=fund_type_filter
        )
    
    def export_positions_csv(self, positions: List[Dict], 
                           filename: Optional[str] = None) -> str:
        """Exportar solo posiciones a CSV"""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"posiciones_{timestamp}.csv"
        
        filepath = os.path.join(EXPORT_DIR, filename)
        
        try:
            df = self._prepare_positions_data(positions)
            if df.empty:
                raise Exception("No hay posiciones para exportar")
            
            df.to_csv(filepath, index=False, encoding='utf-8')
            return filepath
            
        except Exception as e:
            raise Exception(f"Error exportando CSV: {str(e)}")